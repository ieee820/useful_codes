# -*- coding: utf-8 -*-
import sys

from Bio import SeqIO
from glob import glob
from Bio import AlignIO
from Bio.Seq import Seq, reverse_complement
from Bio.SeqRecord import SeqRecord
from Bio.SeqIO.FastaIO import FastaTwoLineWriter
import requests
import os
import pickle
from bs4 import BeautifulSoup
import pandas as pd
import shutil
from Bio.Align.Applications import ClustalwCommandline
import logging
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import easygui as eg

logger = logging.getLogger()
logger.setLevel(logging.INFO)

logfile = os.path.join(os.getcwd(), 'log.txt')
fh = logging.FileHandler(logfile, mode='w')
logger.addHandler(fh)
TH_GENEDTYPE = 3
error_ab1_list = []
id_type = []
gened_type = []

meiqieweidian = {"人源库": ('CCATGGCC', 'TAATAA', 'GCGGCCGC', 'AGCAGA'),
                 "鼠源库": ('GTGCACTT', 'TAATAA', 'GCGGCCGC', 'AGCAGA')}


def igblast_post(nn_fasta_filepath):
    print('do igblast ...')
    url = 'https://www.ncbi.nlm.nih.gov/igblast/igblast.cgi'
    playload = {
        "organism": "human",
        "germline_db_V": "IG_DB/imgt.Homo_sapiens.V.f.orf.p",
        "germline_db_D": "IG_DB/imgt.Homo_sapiens.D.f.orf",
        "germline_db_J": "IG_DB/imgt.Homo_sapiens.J.f.orf",
        "program": "blastn",
        "mismatch_penalty": -1,
        "min_D_match": 5,
        "D_penalty": -2,
        "min_V_length": 9,
        "min_J_length": 0,
        "J_penalty": -2,
        "num_alignments_V": 3,
        "num_alignments_D": 3,
        "num_alignments_J": 3,
        "translation": "true",
        "domain": "imgt",
        "num_clonotype": 100,
        "outfmt": 3,
        "v_focus": "true",
        "num_alignments_additional": 10,
        "evalue": 1,
        "SEARCH_TYPE": "IG",
        "igsource": "new",
        "analyze": "on",
        "CMD": "request"
    }

    files = {
        "queryfile": open(nn_fasta_filepath, "rb")
    }
    response = requests.post(url, data=playload, files=files)

    logger.info('do_igblast')
    logger.info(response.status_code)
    try:
        assert (response.status_code == 200)
    except AssertionError:
        logger.error('do_igblast failed...')
        sys.exit(1)
    return response


def table_to_pd(table):
    rows = []
    for child in table.children:
        row = []
        for td in child:
            try:
                row.append(td.text.replace('\n', ''))
            except:
                continue
        if len(row) > 0:
            rows.append(row)
    df = pd.DataFrame(rows[1:], columns=rows[0])
    # print(df_sum)
    return df


def df_unfold(df_detail):
    res = pd.DataFrame(columns=['ID', 'Count', 'Clone_name'])
    for index, row in df_detail.iterrows():
        # print(index, row['ID'], row['Count'], row['All query name(s)'])
        clone_names = row['All query name(s)']
        # print(clone_names)
        clone_rep_list = clone_names.split(',')
        if len(clone_rep_list) > 1:
            for clone_name in clone_names.split(','):
                # print(clone_name)
                res = res.append({'ID': row['ID'], 'Count': row['Count'], 'Clone_name': clone_name}, ignore_index=True)
        else:
            res = res.append({'ID': row['ID'], 'Count': row['Count'], 'Clone_name': clone_names}, ignore_index=True)

    return res


def parse_igblast_result(response):
    # print(response.text)
    soup = BeautifulSoup(response.text, 'html.parser')
    tables = soup.find_all('table', attrs={'border': 1})
    # print(table)
    for table in tables:
        # print(table.previous.string)
        previous_element = table.previous.string
        if 'Clonotype summary' in previous_element:
            # table to pd
            df_sum = table_to_pd(table)

        elif 'grouped by clonotypes' in previous_element:
            # print(table)
            df_detail = table_to_pd(table)

    # print(df_sum, df_detail)
    res_rep = df_unfold(df_detail)
    res_merge = res_rep.merge(df_sum, on='ID', how='left')
    res_merge['Clone_name_abv'] = res_merge.apply(lambda row: row['Clone_name'].split('_')[0], axis=1)
    Clone_name_abv = res_merge['Clone_name_abv']
    res_merge = res_merge.drop(columns=['Clone_name_abv'])
    res_merge.insert(loc=1, column='Clone_name_abv', value=Clone_name_abv)
    # print(res_merge)
    # res_merge.to_excel('output.xlsx')
    return res_merge


def parse_abysis(response):
    soup = BeautifulSoup(response.text, 'html.parser')
    # forms = soup.find_all('form', attrs={'action': ''})
    # for form in forms:
    #     # print(form)
    #     for child in form.children:
    #         print(child)
    # divs = soup.find_all('div', attrs={'class': 'sequence-display'})
    divs = soup.findAll("div", {"id": re.compile('selection_abm')})
    assert len(divs), 1
    talbes = divs[0].findAll("table", attrs={'class': "results"})
    assert len(talbes), 1
    # print(divs[0].text)
    pd_cdr = table_to_pd(talbes[0])
    # print(pd)
    return pd_cdr


def step2(nn_fasta_filepath, output_xls):
    response = igblast_post(nn_fasta_filepath)
    res_df = parse_igblast_result(response)
    res_df.to_excel(output_xls)


def seq_test():
    for seq_record in SeqIO.parse("F-99-NN.fas", "fasta"):
        print(seq_record.id)
        # print(seq_record.seq)
        print(len(seq_record))


def read_abi():
    handle = open("F/SHS001-570_Seq-F_TSS20210711-021-03705_C01.ab1", "rb")
    for record in SeqIO.parse(handle, "abi"):
        print(record.seq)


def do_abysis(seq_r):
    print('do abysis ...')
    query_seq = str(seq_r.seq)
    url = 'http://www.abysis.org/abysis/sequence_input/key_annotation/key_annotation.cgi'
    playload = {
        "aa_sequence": query_seq,
        "nuc_sequence": '',
        "translation": "sixft",
        "humanorganism": "on"
    }
    response = requests.get(url, params=playload)
    # file_r = open('response_cdr_F.pkl', 'wb')
    # pickle.dump(response, file_r, pickle.HIGHEST_PROTOCOL)
    # file_r.close()
    logger.info('do_abysis')
    logger.info(response.status_code)
    try:
        assert (response.status_code == 200)
    except AssertionError:
        logger.error('do_abysis failed...')
        sys.exit(1)
    return response


def cut_cdr(seq_str, cdr1, cdr2, cdr3):
    res_str1 = seq_str[int(cdr1[0]) - 1: int(cdr1[1])]
    res_str2 = seq_str[int(cdr2[0]) - 1: int(cdr2[1])]
    res_str3 = seq_str[int(cdr3[0]) - 1: int(cdr3[1])]

    return (res_str1, res_str2, res_str3)


# res_str1 = (cdr1, cdr2, cdr3)
# diff threshold > 3
def compare_cdr(res_str1, res_str2):
    diff = 0
    for res1, res2 in zip(res_str1, res_str2):
        # print(res1, res2)
        for x, y in zip(res1, res2):
            if x != y:
                diff += 1

    return diff


def align_to_AA_map(seq_r, cdr3):
    map_dir = {}
    new_cdr3 = cdr3
    seq_align = str(seq_r.seq)
    index = 0
    cnt = 0
    # last = ''
    for s in seq_align:
        if s == '-':
            map_dir[index] = cnt
            # index += 1
            # continue
        else:
            map_dir[index] = cnt
            cnt += 1
        index += 1

    # print(map_dir) res_str3 = seq_str[int(cdr3[0]) - 1: int(cdr3[1])]
    # cut '-' , get new index in aligned AA seq
    for k, v in map_dir.items():
        if v == int(cdr3[0]):
            # print(k)
            new_cdr3[0] = k
            continue
        if v == int(cdr3[1]):
            # print(k)
            new_cdr3[1] = k
            break

    return new_cdr3


# 20 AA: ARDCQEHIGNLKMFPSTWYV
def do_match(cdr_sq, cdr_type):
    # do remove '-'
    cdr_sq = cdr_sq.replace('-', '')
    matchObj_1 = re.search(r'N[ARDCQEHIGNLKMFSTWYV][TS]', cdr_sq)
    matchObj_2 = re.search(r'C', cdr_sq)
    if matchObj_1:
        return cdr_type
    elif matchObj_2:
        return cdr_type
    else:
        return ''


def do_match_ptm(cdrs):
    cdr1 = do_match(cdrs[0], 'cdr1')
    cdr2 = do_match(cdrs[1], 'cdr2')
    cdr3 = do_match(cdrs[2], 'cdr3')

    return cdr1 + '-' + cdr2 + '-' + cdr3


# F_or_R = 'F' or 'R'
def read_aln_get_geneDtype(aln_file, F_or_R):
    genedtype_df = pd.DataFrame(columns=['Clone_name', 'Clone_name_abv', 'GeneDtype', 'PTM'])
    align = AlignIO.read(aln_file, "clustal")
    seq_r = align[0]

    # response = read_pkl("response_cdr_F.pkl")
    response = do_abysis(seq_r)
    pd_cdr = parse_abysis(response)
    print(pd_cdr)
    cdr1 = pd_cdr.iloc[1]['Residues'].split(' - ')
    cdr2 = pd_cdr.iloc[3]['Residues'].split(' - ')
    cdr3 = pd_cdr.iloc[5]['Residues'].split(' - ')
    cdr3 = align_to_AA_map(seq_r, cdr3)
    index = 0
    last_genedtype = 1
    for record in align:
        # print(record.id, record.seq)
        if index < len(align) - 1:
            seq_str1 = str(record.seq)
            seq_id1 = record.id
            res_str1 = cut_cdr(seq_str1, cdr1, cdr2, cdr3)
            # print(align[index + 1])
            seq_str2 = str(align[index + 1].seq)
            seq_id2 = align[index + 1].id
            res_str2 = cut_cdr(seq_str2, cdr1, cdr2, cdr3)

            if index <= 0:
                # print('index == 0')
                # do ptm scan
                ptm_re = do_match_ptm(res_str1)
                genedtype_df = genedtype_df.append(
                    {'Clone_name': seq_id1, 'Clone_name_abv': seq_id1.split('_')[0], 'GeneDtype': last_genedtype,
                     'PTM': ptm_re}, ignore_index=True)

            diff = compare_cdr(res_str1, res_str2)
            # print(seq_id1, seq_id2)
            logger.info(seq_id1)
            logger.info(seq_id2)
            logger.info(res_str1)
            logger.info(res_str2)
            logger.info(diff)
            # print(res_str1, res_str2, diff)

            # do ptm scan
            ptm_re = do_match_ptm(res_str2)
            if diff > TH_GENEDTYPE:
                last_genedtype += 1
                genedtype_df = genedtype_df.append(
                    {'Clone_name': seq_id2, 'Clone_name_abv': seq_id2.split('_')[0], 'GeneDtype': last_genedtype,
                     'PTM': ptm_re}, ignore_index=True)
            else:
                genedtype_df = genedtype_df.append(
                    {'Clone_name': seq_id2, 'Clone_name_abv': seq_id2.split('_')[0], 'GeneDtype': last_genedtype,
                     'PTM': ptm_re}, ignore_index=True)

            index += 1
        else:
            continue

    # rename_
    if F_or_R == 'F':
        genedtype_df.columns = ["Clone_name", "Clone_name_abv", "F-GeneD", "F-PTM"]
        genedtype_df = genedtype_df[["Clone_name_abv", "F-GeneD", "F-PTM"]]
    elif F_or_R == 'R':
        genedtype_df.columns = ["Clone_name", "Clone_name_abv", "R-GeneD", "R-PTM"]
        genedtype_df = genedtype_df[["Clone_name_abv", "R-GeneD", "R-PTM"]]

    return genedtype_df


def dna_enzyme_cutting(seq, cutting_site_left, cutting_site_right, ab1_filename):
    seq_str = str(seq)
    # print(seq_str)
    if (cutting_site_left in seq_str) and (cutting_site_right in seq_str):
        # cut left seq
        pos_left = seq_str.find(cutting_site_left)
        step_left = len(cutting_site_left)
        cut_range_left = pos_left + step_left
        left_seq_str = seq_str[cut_range_left:]
        # print(new_seq_str)
        # cut right seq
        pos_right = left_seq_str.find(cutting_site_right)
        # step_right = len(cutting_site_right)
        # cut_range_right = pos_right + step_right
        right_seq_str = left_seq_str[0:pos_right]
    else:
        error_ab1_list.append(ab1_filename)
        right_seq_str = ''

    return right_seq_str


def abi_to_fasta(ab1_floder_path, cutting_site_left, cutting_site_right, output_nn_path):
    all_sequences = []
    for filename in sorted(glob(ab1_floder_path)):
        handle = open(filename, "rb")
        for record in SeqIO.parse(handle, "abi"):
            seq_cut = dna_enzyme_cutting(record.seq, cutting_site_left, cutting_site_right, record.name)
            if seq_cut != '':
                simple_seq = Seq(seq_cut)
                simple_seq_r = SeqRecord(simple_seq)
                simple_seq_r.id = record.name
                simple_seq_r.description = ''
                all_sequences.append(simple_seq_r)

    # SeqIO.write(all_sequences, "all_sequences.fasta", "fasta")
    handle = open(output_nn_path, 'w')
    writer = FastaTwoLineWriter(handle)
    writer.write_file(all_sequences)
    handle.close()


def dna_to_protein(nn_fasta_file, output_aa_path, if_reverse_complement):
    all_aa_seqs = []
    for seq_record in SeqIO.parse(nn_fasta_file, 'fasta'):
        # print(seq_record.id)
        # print(repr(seq_record.seq))
        seq = seq_record.seq
        if if_reverse_complement:
            # seq = seq.complement()
            seq = reverse_complement(seq)
        rna = seq.transcribe()
        aa_seq = rna.translate()
        # print(protein)
        simple_seq = Seq(aa_seq)
        simple_seq_r = SeqRecord(simple_seq)
        simple_seq_r.id = seq_record.name
        simple_seq_r.description = ''
        all_aa_seqs.append(simple_seq_r)

    handle = open(output_aa_path, 'w')
    writer = FastaTwoLineWriter(handle)
    writer.write_file(all_aa_seqs)
    handle.close()


# F_cuat = ['CCATGGCC' , 'TAATAA']
def step1(F_ab1_folder, R_ab1_folder, F_cut, R_cut, work_dir):
    F_NN_path = os.path.join(work_dir, 'F_NN.fasta')
    R_NN_path = os.path.join(work_dir, 'R_NN.fasta')
    F_AA_path = os.path.join(work_dir, 'F_AA.fasta')
    R_AA_path = os.path.join(work_dir, 'R_AA.fasta')
    abi_to_fasta(F_ab1_folder, F_cut[0], F_cut[1], F_NN_path)
    abi_to_fasta(R_ab1_folder, R_cut[0], R_cut[1], R_NN_path)
    dna_to_protein(F_NN_path, F_AA_path, False)
    dna_to_protein(R_NN_path, R_AA_path, True)


def merge_F_R(F_igblast_path, R_igblast_path, merge_igblast_path):
    F_df = pd.read_excel(F_igblast_path, 'Sheet1')
    R_df = pd.read_excel(R_igblast_path, 'Sheet1')
    F_mini_df = F_df[["ID", "Clone_name_abv", "V gene"]]
    F_mini_df.columns = ["ID-F", "Clone_name_abv", "F-V gene"]

    R_mini_df = R_df[["ID", "Clone_name_abv", "V gene"]]
    R_mini_df.columns = ["ID-R", "Clone_name_abv", "R-V gene"]
    # R_mini_df["ID-R"] = pd.astype(R_mini_df["ID-R"])
    # R_mini_df.astype('object').dtypes

    merge_df = F_mini_df.merge(R_mini_df, on='Clone_name_abv', how='inner')
    # merge_df = merge_df.astype(object)
    merge_df['ID-type'] = merge_df.apply(lambda row: str(row['ID-F']) + '-' + str(row['ID-R']), axis=1)
    # print(merge_df.dtypes)
    # merge_df.astype('object').dtypes
    # merge_df = merge_df.astype(object)
    merge_df = merge_df[["Clone_name_abv", "ID-F", "ID-R", 'ID-type', "F-V gene", "R-V gene"]]
    merge_df.to_excel(merge_igblast_path)


def do_clustalw(clustalw_exe, input_file, output_file):
    print('do clustalw ...')
    clustalw_cline = ClustalwCommandline(clustalw_exe, infile=input_file, outfile=output_file)
    assert os.path.isfile(clustalw_exe), "Clustal W executable missing"
    stdout, stderr = clustalw_cline()


def read_pkl(file_path):
    with open(file_path, 'rb') as inp:
        response = pickle.load(inp)
    return response


def merge_geneDtype(F_df, R_df):
    merge_genedtype_df = F_df.merge(R_df, on='Clone_name_abv', how='inner')
    merge_genedtype_df['GeneD-Type'] = merge_genedtype_df.apply(
        lambda row: str(row['F-GeneD']) + '-' + str(row['R-GeneD']), axis=1)

    merge_genedtype_df = merge_genedtype_df[["Clone_name_abv", "F-GeneD", "R-GeneD", "GeneD-Type", "F-PTM", "R-PTM"]]
    return merge_genedtype_df


def final_merge(igblast, genedtype):
    finale_df = igblast.merge(genedtype, on='Clone_name_abv', how='inner')
    finale_df = finale_df[
        ["Clone_name_abv", "ID-F", "ID-R", "ID-type", "F-V gene", "R-V gene", "F-GeneD", "R-GeneD", "GeneD-Type",
         "F-PTM", "R-PTM"]]
    return finale_df


def id_type_color(df_cell):
    color = 'white'
    if df_cell in id_type:
        color = 'yellow'
    # % color background-color: 'yellow',[color: red, background-color: yellow]
    return 'background-color: %s' % color


def gened_type_color(df_cell):
    color = 'black'
    if df_cell in gened_type:
        color = 'red'

    return 'color: %s' % color


# ['Clone_name_abv','ID-type'] ,final_xls.iloc[48,1]
def apply_color_to_cell(final_xls):
    df_agg = final_xls.groupby(final_xls['ID-type'], sort=True)
    for k, v in df_agg.groups.items():
        s = final_xls.iloc[v, 0].sort_values()
        id_type.append(s.iloc[0])

    df_gene = final_xls.groupby(final_xls['GeneD-Type'], sort=True)
    for k, v in df_gene.groups.items():
        s = final_xls.iloc[v, 0].sort_values()
        gened_type.append(s.iloc[0])

    # apply color to cells
    final_xls = final_xls.style.applymap(id_type_color, subset=['Clone_name_abv']).applymap(gened_type_color, subset=[
        'Clone_name_abv'])

    return final_xls


def init():
    # check if pwd path contains non-English chars
    p_path = os.getcwd()
    regex_ch = r".*?([\u4E00-\u9FA5])"
    regex_space = r"[\t\n\r\f\s]"
    match_ch = re.search(regex_ch, p_path)
    match_sp = re.search(regex_space, p_path)
    if match_ch:
        # print(match_obj.group(1))
        eg.msgbox("检测到程序目录中(或者所在文件夹路径)有中文或者特殊字符,程序将发生异常并退出,请将程序移动到纯英文路径(不能有空格字符)再启动", ok_button="OK")
        sys.exit(1)

    if match_sp:
        # print(match_obj.group(1))
        eg.msgbox("检测到程序目录中(或者所在文件夹路径)有中文或者特殊字符,程序将发生异常并退出,请将程序移动到纯英文路径(不能有空格字符)再启动", ok_button="OK")
        sys.exit(1)

    # select meiqie weidian
    reply = eg.choicebox("请选择酶切位点库类型", choices=["人源库", "鼠源库"])
    if reply is None:
        sys.exit(0)
    # print(reply)
    meiqie = meiqieweidian[reply]
    meiqie_F_L = meiqie[0]
    meiqie_F_R = meiqie[1]
    meiqie_R_L = meiqie[2]
    meiqie_R_R = meiqie[3]

    # check meiqie points
    msg = "轻链(F): " + meiqie_F_L + "(左端)" + meiqie_F_R + "(右端)" + " 重链(R): " + meiqie_R_L + "(左端)" + meiqie_R_R + "(右端)"

    title = "请确认酶切位点是否正确?"
    if eg.ccbox(msg, title):
        pass  # user chose Continue
    else:  # user chose Cancel
        msg = "请输入酶切位点"
        title = "输入酶切位点"
        fieldNames = ["轻链左端(F)", "轻链右端(F)", "重链左端(R)", "重链右端(R)"]
        fieldValues = []  # we start with blanks for the values
        fieldValues = eg.multenterbox(msg, title, fieldNames)

        if fieldValues is None:
            sys.exit(0)
        else:
            meiqie_F_L = fieldValues[0]
            meiqie_F_R = fieldValues[1]
            meiqie_R_L = fieldValues[2]
            meiqie_R_R = fieldValues[3]

    # get ab1 folders
    msg = '请选择文件F链序列所在的文件夹, 序列文件类型必须为ab1后缀'
    title = 'F链序列文件夹(*.ab1)'
    F_ab1_path = eg.diropenbox(msg, title)
    if F_ab1_path is None:
        sys.exit(0)

    msg = '请选择文件R链序列所在的文件夹, 序列文件类型必须为ab1后缀'
    title = 'R链序列文件夹(*.ab1)'
    R_ab1_path = eg.diropenbox(msg, title)
    if R_ab1_path is None:
        sys.exit(0)

    msg = '请选择结果输出的目录'
    title = '序列处理结果会存放到该目录下'
    out_dir = eg.diropenbox(msg, title)
    if out_dir is None:
        sys.exit(0)

    # confirm
    msg = "序列文件夹:" + F_ab1_path + ";" + R_ab1_path
    title = "信息确认"
    if eg.ccbox(msg, title):  # show a Continue/Cancel dialog
        pass  # user chose Continue
    else:  # user chose Cancel
        sys.exit(0)

    msg = "ab1序列文件的命名中应该有下划线作为分隔符,\n例如SHS001-17_Seq-F_TSS20210711-021-03678_E09.ab1中SHS001-17后面字符应该是'_'\n否则程序执行将发生错误"
    title = "信息确认"
    if eg.ccbox(msg, title):  # show a Continue/Cancel dialog
        pass  # user chose Continue
    else:  # user chose Cancel
        sys.exit(0)



    ouput_file_path = os.path.join(out_dir, 'wangyiku_result.xlsx')

    clean_dir = True
    pwd = os.getcwd()
    work_dir = os.path.join(pwd, 'working')
    if not os.path.exists(work_dir):
        os.mkdir(work_dir)
    else:
        if clean_dir:
            shutil.rmtree(work_dir)
            os.mkdir(work_dir)

    F_igblast_path = os.path.join(work_dir, 'F_igblast.xlsx')
    R_igblast_path = os.path.join(work_dir, 'R_igblast.xlsx')
    merge_igblast_path = os.path.join(work_dir, 'merge_igblast.xlsx')
    F_NN_path = os.path.join(work_dir, 'F_NN.fasta')
    R_NN_path = os.path.join(work_dir, 'R_NN.fasta')
    F_AA_path = os.path.join(work_dir, 'F_AA.fasta')
    R_AA_path = os.path.join(work_dir, 'R_AA.fasta')
    F_AA_aln_path = os.path.join(work_dir, 'F_AA_aln.aln')
    R_AA_aln_path = os.path.join(work_dir, 'R_AA_aln.aln')
    # final_xls = os.path.join(work_dir, 'F&R_result.xlsx')
    # step1 ab1 -> nn.fasta; nn -> aa.protein.fasta
    F_ab1_path = os.path.join(F_ab1_path, '*.ab1')
    R_ab1_path = os.path.join(R_ab1_path, '*.ab1')
    step1(F_ab1_path, R_ab1_path, [meiqie_F_L, meiqie_F_R], [meiqie_R_L, meiqie_R_R], work_dir)
    # do igblast , then get id-type
    step2(F_NN_path, F_igblast_path)
    time.sleep(2)
    step2(R_NN_path, R_igblast_path)
    # merge F and R igblast
    merge_F_R(F_igblast_path, R_igblast_path, merge_igblast_path)
    # do clustalw to align F and R AA seqs
    clustalw_exe = os.path.join(pwd, 'clustalw2.exe')
    do_clustalw(clustalw_exe, F_AA_path, F_AA_aln_path)
    do_clustalw(clustalw_exe, R_AA_path, R_AA_aln_path)
    # do abysis to get geneDtype
    F_df = read_aln_get_geneDtype(F_AA_aln_path, "F")
    time.sleep(2)
    R_df = read_aln_get_geneDtype(R_AA_aln_path, "R")
    merge_genedtype_df = merge_geneDtype(F_df, R_df)
    merge_igblast_df = pd.read_excel(merge_igblast_path, 'Sheet1')
    final_df = final_merge(merge_igblast_df, merge_genedtype_df)
    final_df = apply_color_to_cell(final_df)
    final_df.to_excel(ouput_file_path)
    wb = load_workbook(ouput_file_path)
    ws = wb.active
    f1 = ws['E1']  # K1
    k1 = ws['J1']
    yellow = PatternFill("solid", start_color='FFFF00')
    red = PatternFill("solid", fgColor='FF0000')
    f1.fill = yellow
    k1.fill = red
    wb.save(ouput_file_path)
    eg.msgbox("序列处理完毕,输出结果在: "+ ouput_file_path, ok_button="OK")
    if error_ab1_list is not None:
        meiqie_error_path = os.path.join(out_dir, 'meiqie_error.txt')
        with open(meiqie_error_path, 'w') as f:
            for item in error_ab1_list:
                f.write("%s\n" % item)
        eg.msgbox("发现酶切位点错误的序列,请查看: " + meiqie_error_path, ok_button="OK")



if __name__ == "__main__":
    init()
    # print(error_ab1_list)
